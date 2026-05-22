"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         CLASIFICADOR DE LOCATIONS - PEMEX / TRANSPORTE MEXICO              ║
║         Identifica País y Estado a partir de texto libre                   ║
║  Versión: 2.0                                                               ║
╚══════════════════════════════════════════════════════════════════════════════╝

ESTRATEGIA GENERAL:
  1. Normalización (minúsculas, sin acentos, sin puntuación extra)
  2. Búsqueda exacta en diccionario de locations conocidas
  3. Búsqueda de substring (la key más larga presente en el texto)
  4. Extracción de abreviaciones de estado (VER., TAMPS., etc.)
  5. Búsqueda de nombre completo de estado en el texto
  6. Palabras clave PEMEX / carreteras → México (estado=REVISAR)
  7. Fuzzy matching como última línea de defensa
  8. Si confianza baja → "REVISAR"

DEPENDENCIAS:
  pip install rapidfuzz unidecode openpyxl
"""

import re
import unicodedata
from typing import Optional
from rapidfuzz import process, fuzz

# ═══════════════════════════════════════════════════════════════════════
# 1. NORMALIZACIÓN DE TEXTO
# ═══════════════════════════════════════════════════════════════════════

def normalize(text: str) -> str:
    """
    Normaliza para comparación:
    - Minúsculas
    - Sin acentos/tildes
    - Sin puntuación (excepto espacios)
    - Espacios colapsados
    """
    if not text:
        return ""
    text = unicodedata.normalize("NFD", str(text))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


# ═══════════════════════════════════════════════════════════════════════
# 2. DICCIONARIO PRINCIPAL  (key = texto normalizado → (país, estado))
# ═══════════════════════════════════════════════════════════════════════

LOCATION_DICT: dict = {
    # ── CDMX / DF ───────────────────────────────────────────────────────
    "cdmx":                    ("México", "Ciudad de México"),
    "df":                      ("México", "Ciudad de México"),
    "distrito federal":        ("México", "Ciudad de México"),
    "cd de mexico":            ("México", "Ciudad de México"),
    "ciudad de mexico":        ("México", "Ciudad de México"),
    "azcapotzalco":            ("México", "Ciudad de México"),
    "miguel hidalgo":          ("México", "Ciudad de México"),
    "18 de marzo":             ("México", "Ciudad de México"),
    "anil":                    ("México", "Ciudad de México"),
    "anll":                    ("México", "Ciudad de México"),

    # ── REFINERÍAS ───────────────────────────────────────────────────────
    "refineria cadereyta":              ("México", "Nuevo León"),
    "refineria de cadereyta":           ("México", "Nuevo León"),
    "cadereyta":                        ("México", "Nuevo León"),
    "cadereyta jimenez":                ("México", "Nuevo León"),
    "refineria tula":                   ("México", "Hidalgo"),
    "tula de allende":                  ("México", "Hidalgo"),
    "tula":                             ("México", "Hidalgo"),
    "refineria salamanca":              ("México", "Guanajuato"),
    "salamanca":                        ("México", "Guanajuato"),
    "refineria salina cruz":            ("México", "Oaxaca"),
    "salina cruz":                      ("México", "Oaxaca"),
    "refineria minatitlan":             ("México", "Veracruz"),
    "minatitlan":                       ("México", "Veracruz"),
    "refineria madero":                 ("México", "Tamaulipas"),
    "refineria cd madero":              ("México", "Tamaulipas"),
    "refineria ciudad madero":          ("México", "Tamaulipas"),

    # ── TERMINALES Y COMPLEJOS ───────────────────────────────────────────
    "dos bocas":                ("México", "Tabasco"),
    "puerto dos bocas":         ("México", "Tabasco"),
    "pajaritos":                ("México", "Veracruz"),
    "terminal pajaritos":       ("México", "Veracruz"),
    "tasp pajaritos":           ("México", "Veracruz"),
    "cangrejera":               ("México", "Veracruz"),
    "complejo petroquimico cangrejera":    ("México", "Veracruz"),
    "complejo petroquimico cagrefera":     ("México", "Veracruz"),
    "complejo petroquimicos martin tex":   ("México", "Puebla"),
    "bajos de la gallega":      ("México", "Veracruz"),
    "ce bajos de la gallega":   ("México", "Veracruz"),
    "c e bajos de la gallega":  ("México", "Veracruz"),
    "escamela":                 ("México", "Veracruz"),
    "tierra blanca":            ("México", "Veracruz"),
    "cpq san martin texmelucan": ("México", "Puebla"),
    "complejo petroquimico san martin texmelucan": ("México", "Puebla"),
    "san martin texmelucan":    ("México", "Puebla"),
    "complejo san martin texm": ("México", "Puebla"),
    "c p san martin texmelucan": ("México", "Puebla"),
    "complejo procesador de gas reynosa burgos": ("México", "Tamaulipas"),
    "complejo petroquimico pajaritos": ("México", "Veracruz"),

    # ── PLATAFORMAS / REGION MARINA ─────────────────────────────────────
    "akal c":                   ("México", "Campeche"),
    "akal-c":                   ("México", "Campeche"),
    "nohoch a":                 ("México", "Campeche"),
    "nohoch-a":                 ("México", "Campeche"),
    "cantarell":                ("México", "Campeche"),
    "activo de produccion cantarell": ("México", "Campeche"),
    "region marina noroeste":   ("México", "Campeche"),
    "region marina noreste":    ("México", "Campeche"),
    "region marina suroeste":   ("México", "Campeche"),
    "region marina sw":         ("México", "Campeche"),
    "marine region sw":         ("México", "Campeche"),
    "marine region nw":         ("México", "Campeche"),
    "marino noroeste":          ("México", "Campeche"),
    "marino noreste":           ("México", "Campeche"),
    "chalan pemex 538":         ("México", "Campeche"),
    "pmx offshore":             ("México", "Campeche"),
    "p m x offshore":           ("México", "Campeche"),

    # ── TAD (Terminales Almacenamiento y Distribución) ────────────────────
    "tad acapulco":             ("México", "Guerrero"),
    "tad aguascalientes":       ("México", "Aguascalientes"),
    "tad campeche":             ("México", "Campeche"),
    "tad cancun":               ("México", "Quintana Roo"),
    "tad cadereyta":            ("México", "Nuevo León"),
    "tad cd obregon":           ("México", "Sonora"),
    "tad cd victoria":          ("México", "Tamaulipas"),
    "tad ciudad victoria":      ("México", "Tamaulipas"),
    "tad cd juarez":            ("México", "Chihuahua"),
    "tad ciudad juarez":        ("México", "Chihuahua"),
    "tad ciudad madero":        ("México", "Tamaulipas"),
    "tad chihuahua":            ("México", "Chihuahua"),
    "tad coatzacoalcos":        ("México", "Veracruz"),
    "tad cuernavaca":           ("México", "Morelos"),
    "tad cuautla":              ("México", "Morelos"),
    "tad durango":              ("México", "Durango"),
    "tad el castillo":          ("México", "Jalisco"),
    "tad escamela":             ("México", "Veracruz"),
    "tad gomez palacio":        ("México", "Durango"),
    "tad guadalajara":          ("México", "Jalisco"),
    "tad guaymas":              ("México", "Sonora"),
    "tad hermosillo":           ("México", "Sonora"),
    "tad hidalgo del parral":   ("México", "Chihuahua"),
    "tad iguala":               ("México", "Guerrero"),
    "tad irapuato":             ("México", "Guanajuato"),
    "tad lazaro cardenas":      ("México", "Michoacán"),
    "tad leon":                 ("México", "Guanajuato"),
    "tad madero":               ("México", "Tamaulipas"),
    "tad magdalena":            ("México", "Sonora"),
    "tad manzanillo":           ("México", "Colima"),
    "tad matehuala":            ("México", "San Luis Potosí"),
    "tad mazatlan":             ("México", "Sinaloa"),
    "tad mexicali":             ("México", "Baja California"),
    "tad miahuatlan":           ("México", "Puebla"),
    "tad minatitlan":           ("México", "Veracruz"),
    "tad monclova":             ("México", "Coahuila"),
    "tad morelia":              ("México", "Michoacán"),
    "tad monterrey":            ("México", "Nuevo León"),
    "tad nogales":              ("México", "Sonora"),
    "tad nuevo laredo":         ("México", "Tamaulipas"),
    "tad oaxaca":               ("México", "Oaxaca"),
    "tad pachuca":              ("México", "Hidalgo"),
    "tad pajaritos":            ("México", "Veracruz"),
    "tad perote":               ("México", "Veracruz"),
    "tad poza rica":            ("México", "Veracruz"),
    "tad progreso":             ("México", "Yucatán"),
    "tad puebla":               ("México", "Puebla"),
    "tad queretaro":            ("México", "Querétaro"),
    "tad reynosa":              ("México", "Tamaulipas"),
    "tad sabinas":              ("México", "Coahuila"),
    "tad salina cruz":          ("México", "Oaxaca"),
    "tad salamanca":            ("México", "Guanajuato"),
    "tad saltillo":             ("México", "Coahuila"),
    "tad san jose iturbide":    ("México", "Guanajuato"),
    "tad san luis potosi":      ("México", "San Luis Potosí"),
    "tad santa catarina":       ("México", "Nuevo León"),
    "tad sta catarina":         ("México", "Nuevo León"),
    "tad tampico":              ("México", "Tamaulipas"),
    "tad tapachula":            ("México", "Chiapas"),
    "tad tepeixtles":           ("México", "Colima"),
    "tad tierra blanca":        ("México", "Veracruz"),
    "tad toluca":               ("México", "Estado de México"),
    "tad topolobampo":          ("México", "Sinaloa"),
    "tad tula":                 ("México", "Hidalgo"),
    "tad tuxtla gutierrez":     ("México", "Chiapas"),
    "tad tuxpan":               ("México", "Veracruz"),
    "tad itzoil tuxpan":        ("México", "Veracruz"),
    "tad uruapan":              ("México", "Michoacán"),
    "tad veracruz":             ("México", "Veracruz"),
    "tad villahermosa":         ("México", "Tabasco"),
    "tad zacatecas":            ("México", "Zacatecas"),
    "tad zamora":               ("México", "Michoacán"),
    "tad zapopan":              ("México", "Jalisco"),
    "tad 18 de marzo":          ("México", "Ciudad de México"),
    "tad 18 de marzo azcapotzalco": ("México", "Ciudad de México"),
    "tad azcapotzalco":         ("México", "Ciudad de México"),

    # ── CIUDADES/MUNICIPIOS MEXICO ────────────────────────────────────────
    "acapulco":              ("México", "Guerrero"),
    "acambay":               ("México", "Estado de México"),
    "acayucan":              ("México", "Veracruz"),
    "acacoyagua":            ("México", "Chiapas"),
    "actopan":               ("México", "Hidalgo"),
    "aguascalientes":        ("México", "Aguascalientes"),
    "altamira":              ("México", "Tamaulipas"),
    "apaseo el grande":      ("México", "Guanajuato"),
    "arriaga":               ("México", "Chiapas"),
    "atlacomulco":           ("México", "Estado de México"),
    "bahia de acapulco":     ("México", "Guerrero"),
    "bahia concepcion":      ("Chile", "Biobío"),
    "bella vista":           ("México", "Michoacán"),
    "cabo san lucas":        ("México", "Baja California Sur"),
    "campeche":              ("México", "Campeche"),
    "cananea":               ("México", "Sonora"),
    "cancun":                ("México", "Quintana Roo"),
    "carmen":                ("México", "Campeche"),
    "cd carmen":             ("México", "Campeche"),
    "cd del carmen":         ("México", "Campeche"),
    "ciudad del carmen":     ("México", "Campeche"),
    "cd madero":             ("México", "Tamaulipas"),
    "ciudad madero":         ("México", "Tamaulipas"),
    "cd juarez":             ("México", "Chihuahua"),
    "ciudad juarez":         ("México", "Chihuahua"),
    "cd victoria":           ("México", "Tamaulipas"),
    "ciudad victoria":       ("México", "Tamaulipas"),
    "cd del maiz":           ("México", "San Luis Potosí"),
    "cd guzman":             ("México", "Jalisco"),
    "cd mendoza":            ("México", "Veracruz"),
    "cd valles":             ("México", "San Luis Potosí"),
    "celaya":                ("México", "Guanajuato"),
    "centro":                ("México", "Tabasco"),
    "chiapas":               ("México", "Chiapas"),
    "chihuahua":             ("México", "Chihuahua"),
    "coatzacoalcos":         ("México", "Veracruz"),
    "concepcion del oro":    ("México", "Zacatecas"),
    "concha del oro":        ("México", "Zacatecas"),
    "cordoba":               ("México", "Veracruz"),
    "cosamaloapan":          ("México", "Veracruz"),
    "cosio":                 ("México", "Aguascalientes"),
    "cosoleacaque":          ("México", "Veracruz"),
    "cosoloacaque":          ("México", "Veracruz"),
    "cotaxtla":              ("México", "Veracruz"),
    "cozumel":               ("México", "Quintana Roo"),
    "cuautla":               ("México", "Morelos"),
    "cuernavaca":            ("México", "Morelos"),
    "culiacan":              ("México", "Sinaloa"),
    "durango":               ("México", "Durango"),
    "encarnacion diaz":      ("México", "Jalisco"),
    "fresnillo":             ("México", "Zacatecas"),
    "gomez palacio":         ("México", "Durango"),
    "guadalajara":           ("México", "Jalisco"),
    "guanajuato":            ("México", "Guanajuato"),
    "guaymas":               ("México", "Sonora"),
    "hermosillo":            ("México", "Sonora"),
    "hidalgo del parral":    ("México", "Chihuahua"),
    "huixtla":               ("México", "Chiapas"),
    "iguala":                ("México", "Guerrero"),
    "imuris":                ("México", "Sonora"),
    "irapuato":              ("México", "Guanajuato"),
    "ixtepec":               ("México", "Oaxaca"),
    "jalapa":                ("México", "Veracruz"),
    "xalapa":                ("México", "Veracruz"),
    "kantunil":              ("México", "Yucatán"),
    "lagos de moreno":       ("México", "Jalisco"),
    "las choapas":           ("México", "Veracruz"),
    "lazaro cardenas":       ("México", "Michoacán"),
    "leon":                  ("México", "Guanajuato"),
    "linares":               ("México", "Nuevo León"),
    "los mochis":            ("México", "Sinaloa"),
    "manzanillo":            ("México", "Colima"),
    "matamoros":             ("México", "Tamaulipas"),
    "matehuala":             ("México", "San Luis Potosí"),
    "mazatlan":              ("México", "Sinaloa"),
    "merida":                ("México", "Yucatán"),
    "mexicali":              ("México", "Baja California"),
    "miahuatlan":            ("México", "Puebla"),
    "misantla":              ("México", "Veracruz"),
    "monclova":              ("México", "Coahuila"),
    "monterrey":             ("México", "Nuevo León"),
    "morelia":               ("México", "Michoacán"),
    "morelos":               ("México", "Morelos"),
    "navojoa":               ("México", "Sonora"),
    "nogales":               ("México", "Sonora"),
    "nuevo laredo":          ("México", "Tamaulipas"),
    "oaxaca":                ("México", "Oaxaca"),
    "ocozocoautla":          ("México", "Chiapas"),
    "ocozocuautla":          ("México", "Chiapas"),
    "pachuca":               ("México", "Hidalgo"),
    "palenque":              ("México", "Chiapas"),
    "palmillas":             ("México", "Querétaro"),
    "patzcuaro":             ("México", "Michoacán"),
    "perote":                ("México", "Veracruz"),
    "pinotepa nacional":     ("México", "Oaxaca"),
    "progreso":              ("México", "Yucatán"),
    "puebla":                ("México", "Puebla"),
    "puerto madero":         ("México", "Chiapas"),
    "puerto san benito":     ("México", "Chiapas"),
    "puerto vallarta":       ("México", "Jalisco"),
    "queretaro":             ("México", "Querétaro"),
    "reynosa":               ("México", "Tamaulipas"),
    "sabinas":               ("México", "Coahuila"),
    "saltillo":              ("México", "Coahuila"),
    "san jose del cabo":     ("México", "Baja California Sur"),
    "san jose iturbide":     ("México", "Guanajuato"),
    "san juan de los lagos":("México", "Jalisco"),
    "san juan del rio":      ("México", "Querétaro"),
    "san luis potosi":       ("México", "San Luis Potosí"),
    "san marcos":            ("México", "Guerrero"),
    "san miguel de allende": ("México", "Guanajuato"),
    "santa catarina":        ("México", "Nuevo León"),
    "sayula de aleman":      ("México", "Veracruz"),
    "tampico":               ("México", "Tamaulipas"),
    "tapachula":             ("México", "Chiapas"),
    "tecate":                ("México", "Baja California"),
    "tehuacan":              ("México", "Puebla"),
    "tehuantepec":           ("México", "Oaxaca"),
    "tepic":                 ("México", "Nayarit"),
    "teziutlan":             ("México", "Puebla"),
    "tijuana":               ("México", "Baja California"),
    "tlaxcala":              ("México", "Tlaxcala"),
    "toluca":                ("México", "Estado de México"),
    "topolobampo":           ("México", "Sinaloa"),
    "torreon":               ("México", "Coahuila"),
    "tuxtla gutierrez":      ("México", "Chiapas"),
    "tulancingo":            ("México", "Hidalgo"),
    "tuxpan":                ("México", "Veracruz"),
    "uruapan":               ("México", "Michoacán"),
    "veracruz":              ("México", "Veracruz"),
    "villa aldama":          ("México", "Veracruz"),
    "villahermosa":          ("México", "Tabasco"),
    "zacatecas":             ("México", "Zacatecas"),
    "zamora":                ("México", "Michoacán"),
    "zapopan":               ("México", "Jalisco"),
    "zihuatanejo":           ("México", "Guerrero"),
    "aeropuerto de san jose del cabo": ("México", "Baja California Sur"),


    # ── ADICIONALES DETECTADOS EN DATOS REALES ───────────────────────
    "rosarito":              ("México",        "Baja California"),
    "rosarito monoboya":     ("México",        "Baja California"),
    "mexico df":             ("México",        "Ciudad de México"),
    "la paz":                ("México",        "Baja California Sur"),
    "tad la paz":            ("México",        "Baja California Sur"),
    "pacifico norte":        ("México",        "Sinaloa"),
    "lerma":                 ("México",        "Campeche"),
    "puerto de lerma":       ("México",        "Campeche"),
    "terminal maritima madero": ("México",     "Tamaulipas"),
    "tasp madero":           ("México",        "Tamaulipas"),
    "madero muelle 2":       ("México",        "Tamaulipas"),
    "terminal madero":       ("México",        "Tamaulipas"),
    "transpeninsular":       ("México",        "Baja California Sur"),
    "los barriles":          ("México",        "Baja California Sur"),
    "chapalilla":            ("México",        "Nayarit"),
    "compostela":            ("México",        "Nayarit"),
    "plataforma akal":       ("México",        "Campeche"),
    "akal":                  ("México",        "Campeche"),
    "tad castillo":          ("México",        "Jalisco"),
    "tad 18 marzo":          ("México",        "Ciudad de México"),
    "estacion de bombeo linares": ("México",   "Nuevo León"),
    "piedras negras":        ("México",        "Coahuila"),
    "san diego":             ("EUA",           "California"),
    "port arthur":           ("EUA",           "Texas"),
    "louisiana":             ("EUA",           "Luisiana"),
    "lake charles":          ("EUA",           "Luisiana"),
    "pasadena":              ("EUA",           "Texas"),
    "paulsboro":             ("EUA",           "Nueva Jersey"),
    "nueva jersey":          ("EUA",           "Nueva Jersey"),
    "singapore":             ("Singapur",      ""),
    "singapur":              ("Singapur",      ""),
    "indonesia":             ("Indonesia",     ""),
    "nipah":                 ("Indonesia",     ""),
    "canada":                ("Canadá",        ""),
    "halifax":               ("Canadá",        "Nueva Escocia"),
    "amsterdam":             ("Países Bajos",  ""),
    "athens":                ("Grecia",        ""),
    "grecia":                ("Grecia",        ""),

    # ── INTERNACIONAL ─────────────────────────────────────────────────────
    "aruba":        ("Aruba",     ""),
    "amberes":      ("Bélgica",   "Amberes"),
    "belgica":      ("Bélgica",   ""),
    "balboa":       ("Panamá",    ""),
    "panama":       ("Panamá",    ""),
    "colon panama": ("Panamá",    "Colón"),
    "talcahuano":   ("Chile",     "Biobío"),
    "chile":        ("Chile",     ""),
    "beaumont":     ("EUA",       "Texas"),
    "texas":        ("EUA",       "Texas"),
    "bolivar roads":("EUA",       "Texas"),
    "galveston":    ("EUA",       "Texas"),
    "houston":      ("EUA",       "Texas"),
    "california":   ("EUA",       "California"),
    "brownsville":  ("EUA",       "Texas"),
    "alemania":     ("Alemania",  ""),
    "eeuu":         ("EUA",       ""),
}


# ═══════════════════════════════════════════════════════════════════════
# 3. ABREVIACIONES DE ESTADOS
# ═══════════════════════════════════════════════════════════════════════

STATE_ABBR: dict = {
    "ags":         "Aguascalientes",
    "bc":          "Baja California",
    "bcn":         "Baja California",
    "bcs":         "Baja California Sur",
    "camp":        "Campeche",
    "chis":        "Chiapas",
    "chih":        "Chihuahua",
    "ch":          "Chihuahua",
    "coah":        "Coahuila",
    "col":         "Colima",
    "cdmx":        "Ciudad de México",
    "df":          "Ciudad de México",
    "dgo":         "Durango",
    "gto":         "Guanajuato",
    "gro":         "Guerrero",
    "hgo":         "Hidalgo",
    "jal":         "Jalisco",
    "mex":         "Estado de México",
    "edomex":      "Estado de México",
    "edo mex":     "Estado de México",
    "mich":        "Michoacán",
    "mor":         "Morelos",
    "nay":         "Nayarit",
    "nl":          "Nuevo León",
    "n l":         "Nuevo León",
    "oax":         "Oaxaca",
    "pue":         "Puebla",
    "qro":         "Querétaro",
    "q roo":       "Quintana Roo",
    "q ro":        "Quintana Roo",
    "qr":          "Quintana Roo",
    "slp":         "San Luis Potosí",
    "s l p":       "San Luis Potosí",
    "sin":         "Sinaloa",
    "son":         "Sonora",
    "tab":         "Tabasco",
    "tamps":       "Tamaulipas",
    "tlax":        "Tlaxcala",
    "ver":         "Veracruz",
    "yuc":         "Yucatán",
    "zac":         "Zacatecas",
    "gtez":        "Chiapas",
}

STATE_NAMES: dict = {
    "aguascalientes":    "Aguascalientes",
    "baja california sur": "Baja California Sur",
    "baja california":   "Baja California",
    "campeche":          "Campeche",
    "chiapas":           "Chiapas",
    "chihuahua":         "Chihuahua",
    "coahuila":          "Coahuila",
    "colima":            "Colima",
    "ciudad de mexico":  "Ciudad de México",
    "durango":           "Durango",
    "guanajuato":        "Guanajuato",
    "guerrero":          "Guerrero",
    "hidalgo":           "Hidalgo",
    "jalisco":           "Jalisco",
    "estado de mexico":  "Estado de México",
    "edo de mexico":     "Estado de México",
    "michoacan":         "Michoacán",
    "morelos":           "Morelos",
    "nayarit":           "Nayarit",
    "nuevo leon":        "Nuevo León",
    "oaxaca":            "Oaxaca",
    "puebla":            "Puebla",
    "queretaro":         "Querétaro",
    "quintana roo":      "Quintana Roo",
    "san luis potosi":   "San Luis Potosí",
    "sinaloa":           "Sinaloa",
    "sonora":            "Sonora",
    "tabasco":           "Tabasco",
    "tamaulipas":        "Tamaulipas",
    "tlaxcala":          "Tlaxcala",
    "veracruz":          "Veracruz",
    "yucatan":           "Yucatán",
    "zacatecas":         "Zacatecas",
}

PEMEX_KEYWORDS = [
    "pemex", "tad ", "tar ", "asa ", "refineria", "refinery",
    "plataforma", "complejo petroquimico", "cpq",
    "terminal maritima", "estacion de bombeo",
    "region marina", "marine region",
    "activo de produccion", "campo petrolero",
]

CARRETERA_KEYWORDS = [
    "carretera", "autopista", "carr ", "carrt", "arco norte",
    "libramiento", "carr.", "autp",
]

TRAYECTO_KEYWORDS = [
    "trayecto", "taryecto", "tayecto", "traslado",
    "de la tad", "de la tar", "hacia la tad", "hacia la tar",
    "durante su", "duarante", "duarnte",
]


# ═══════════════════════════════════════════════════════════════════════
# 4. FUNCIONES AUXILIARES
# ═══════════════════════════════════════════════════════════════════════

def _result(loc, pais, estado, confianza, razon, verbose):
    out = {
        "LOCATION_ORIGINAL": loc,
        "PAIS": pais,
        "ESTADO": estado,
        "CONFIANZA": confianza,
    }
    if verbose:
        out["RAZON"] = razon
    return out


def _extract_state_abbr(norm: str) -> Optional[str]:
    """Busca abreviaciones de estado como tokens completos."""
    sorted_abbrs = sorted(STATE_ABBR.keys(), key=len, reverse=True)
    for abbr in sorted_abbrs:
        pattern = r"(?:^|[\s,]){}(?:$|[\s,\.])".format(re.escape(abbr))
        if re.search(pattern, norm):
            return STATE_ABBR[abbr]
    return None


def _extract_state_full(norm: str) -> Optional[str]:
    """Busca nombre completo de estado en el texto."""
    for state in sorted(STATE_NAMES.keys(), key=len, reverse=True):
        if state in norm:
            return STATE_NAMES[state]
    return None


def _find_best_substring(norm: str, keys: list) -> tuple:
    """
    Busca qué key del diccionario está contenida en el texto.
    Prioriza la key más larga (más específica).
    """
    best_key = None
    best_score = 0
    for key in sorted(keys, key=len, reverse=True):
        if len(key) < 4:
            continue
        if key in norm:
            if len(key) > len(best_key or ""):
                best_key = key
                best_score = 100
    return best_key, best_score, None


def _trayecto_state(norm: str) -> Optional[str]:
    """
    Para textos de trayecto (DE LA TAD X HACIA LA TAD Y),
    extrae el estado de la ubicación origen (primera mención).
    """
    # Buscar primer match de TAD/TAR con estado
    patterns = [
        r"tad\s+([a-z\s]+?)(?:,\s*([a-z\s\.]+?))?(?:\s+(?:hacia|a la|a|al)\s|$)",
        r"tar\s+([a-z\s]+?)(?:,\s*([a-z\s\.]+?))?(?:\s+(?:hacia|a la|a|al)\s|$)",
    ]
    for pat in patterns:
        m = re.search(pat, norm)
        if m:
            ciudad = m.group(1).strip() if m.group(1) else ""
            estado_abr = m.group(2).strip() if m.group(2) else ""
            # Intentar resolver abreviación
            if estado_abr:
                for abbr, est in STATE_ABBR.items():
                    if abbr == estado_abr.replace(" ", "").replace(".", ""):
                        return est
            # Intentar resolver ciudad
            if ciudad in LOCATION_DICT:
                return LOCATION_DICT[ciudad][1]
    return None


# ═══════════════════════════════════════════════════════════════════════
# 5. FUNCIÓN PRINCIPAL DE CLASIFICACIÓN
# ═══════════════════════════════════════════════════════════════════════

def classify_location(
    location: str,
    fuzzy_threshold: int = 80,
    verbose: bool = False,
) -> dict:
    """
    Clasifica una location libre: devuelve país, estado y nivel de confianza.

    Parámetros
    ----------
    location        : Texto libre de ubicación
    fuzzy_threshold : Score mínimo para fuzzy match (0-100)
    verbose         : Incluir campo RAZON en el resultado

    Retorna
    -------
    dict con: LOCATION_ORIGINAL, PAIS, ESTADO, CONFIANZA [, RAZON]
    Confianza posible: "ALTA" | "MEDIA" | "BAJA"
    País/Estado = "REVISAR" cuando la confianza es insuficiente
    """
    if not location or str(location).strip() in ("-", "", "None", "nan"):
        return _result(location, "REVISAR", "REVISAR", "BAJA", "Vacío o nulo", verbose)

    raw = str(location).strip()
    norm = normalize(raw)

    # ── Paso 1: Exacto en diccionario ─────────────────────────────────
    if norm in LOCATION_DICT:
        pais, estado = LOCATION_DICT[norm]
        return _result(raw, pais, estado, "ALTA", "Diccionario exacto", verbose)

    # ── Paso 2: Key del diccionario contenida en el texto ─────────────
    best_key, best_score, _ = _find_best_substring(norm, list(LOCATION_DICT.keys()))
    if best_key:
        pais, estado = LOCATION_DICT[best_key]
        return _result(raw, pais, estado, "ALTA", f"Substring exacto: '{best_key}'", verbose)

    # ── Paso 3: Abreviación de estado ─────────────────────────────────
    estado_abbr = _extract_state_abbr(norm)
    if estado_abbr:
        return _result(raw, "México", estado_abbr, "ALTA", f"Abrev. estado: {estado_abbr}", verbose)

    # ── Paso 4: Nombre completo de estado ─────────────────────────────
    estado_full = _extract_state_full(norm)
    if estado_full:
        return _result(raw, "México", estado_full, "ALTA", f"Estado completo: {estado_full}", verbose)

    # ── Paso 5: Texto de trayecto (DE LA TAD X HACIA LA TAD Y) ────────
    is_trayecto = any(kw in norm for kw in TRAYECTO_KEYWORDS)
    if is_trayecto:
        estado_tray = _trayecto_state(norm)
        if estado_tray:
            return _result(raw, "México", estado_tray, "MEDIA",
                           f"Trayecto - estado origen: {estado_tray}", verbose)
        return _result(raw, "México", "REVISAR", "MEDIA",
                       "Trayecto PEMEX (estado no determinado)", verbose)

    # ── Paso 6: Keywords PEMEX ─────────────────────────────────────────
    for kw in PEMEX_KEYWORDS:
        if kw in norm:
            return _result(raw, "México", "REVISAR", "MEDIA",
                           f"Keyword PEMEX: '{kw}'", verbose)

    # ── Paso 7: Patrones de carretera ─────────────────────────────────
    for kw in CARRETERA_KEYWORDS:
        if kw in norm:
            return _result(raw, "México", "REVISAR", "MEDIA",
                           f"Patrón carretera: '{kw}'", verbose)

    # ── Paso 8: Fuzzy matching sobre diccionario ───────────────────────
    match = process.extractOne(
        norm,
        list(LOCATION_DICT.keys()),
        scorer=fuzz.partial_ratio,
        score_cutoff=fuzzy_threshold,
    )
    if match:
        matched_key, score, _ = match
        pais, estado = LOCATION_DICT[matched_key]
        confidence = "ALTA" if score >= 90 else "MEDIA"
        return _result(raw, pais, estado, confidence,
                       f"Fuzzy: '{matched_key}' ({score}%)", verbose)

    # ── Sin resultado confiable ────────────────────────────────────────
    return _result(raw, "REVISAR", "REVISAR", "BAJA", "Sin coincidencia", verbose)


# ═══════════════════════════════════════════════════════════════════════
# 6. PROCESAMIENTO EN LOTE CON CACHÉ
# ═══════════════════════════════════════════════════════════════════════

def classify_batch(
    locations: list,
    fuzzy_threshold: int = 80,
    verbose: bool = False,
) -> list:
    """
    Clasifica miles de locations eficientemente usando caché.

    Parámetros
    ----------
    locations      : Lista de strings
    fuzzy_threshold: Umbral fuzzy
    verbose        : Incluir RAZON

    Retorna
    -------
    Lista de dicts con resultados
    """
    cache = {}
    results = []
    for loc in locations:
        cache_key = str(loc).strip().upper()
        if cache_key not in cache:
            cache[cache_key] = classify_location(loc, fuzzy_threshold, verbose)
        r = cache[cache_key].copy()
        r["LOCATION_ORIGINAL"] = loc
        results.append(r)
    return results


# ═══════════════════════════════════════════════════════════════════════
# 7. NORMALIZACIÓN DE VARIANTES
# ═══════════════════════════════════════════════════════════════════════

def normalize_variants(locations: list, threshold: int = 85) -> dict:
    """
    Detecta variantes del mismo lugar y propone forma canónica.

    Ejemplo: ["COATZACOALCOS", "Coatzacoalcos", "COATZACOALCOS, VER."]
    → todas → "COATZACOALCOS"

    Retorna
    -------
    dict: {variante: canónico}
    """
    unique = list(set(str(l).strip() for l in locations
                      if l and str(l).strip() not in ("-", "", "None")))
    norm_to_orig = {normalize(u): u for u in unique}
    canonical_map = {}
    used = set()

    for norm_key in sorted(norm_to_orig.keys(), key=len):
        if norm_key in used:
            continue
        canonical = norm_to_orig[norm_key]
        canonical_map[canonical] = canonical
        used.add(norm_key)
        for other_norm, other_orig in norm_to_orig.items():
            if other_norm in used:
                continue
            score = fuzz.token_sort_ratio(norm_key, other_norm)
            if score >= threshold:
                canonical_map[other_orig] = canonical
                used.add(other_norm)
    return canonical_map


# ═══════════════════════════════════════════════════════════════════════
# 8. PROCESAMIENTO DE EXCEL
# ═══════════════════════════════════════════════════════════════════════

def process_excel(
    input_path: str,
    output_path: str,
    location_col: str = "LOCATION",
    sheet_name: str = None,
    fuzzy_threshold: int = 80,
    verbose: bool = True,
) -> None:
    """
    Lee Excel con columna LOCATION, clasifica y guarda con nuevas columnas.

    Las celdas se colorean según confianza:
      Verde  = ALTA
      Amarillo = MEDIA
      Rojo   = BAJA
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = [cell.value for cell in ws[1]]
    try:
        loc_idx = header.index(location_col)
    except ValueError:
        loc_idx = next(
            (i for i, h in enumerate(header)
             if h and normalize(str(h)) == normalize(location_col)),
            None
        )
        if loc_idx is None:
            raise ValueError(f"Columna '{location_col}' no encontrada. Columnas: {header}")

    locations = [row[loc_idx] for row in ws.iter_rows(min_row=2, values_only=True)]

    print(f"  Clasificando {len(locations):,} registros...")
    results = classify_batch(locations, fuzzy_threshold, verbose)

    new_cols = ["PAIS", "ESTADO", "CONFIANZA"]
    if verbose:
        new_cols.append("RAZON")

    start_col = ws.max_column + 1
    for i, col_name in enumerate(new_cols):
        cell = ws.cell(row=1, column=start_col + i, value=col_name)
        cell.font = Font(bold=True)

    fills = {
        "ALTA":  PatternFill("solid", fgColor="C6EFCE"),
        "MEDIA": PatternFill("solid", fgColor="FFEB9C"),
        "BAJA":  PatternFill("solid", fgColor="FFC7CE"),
    }

    for row_idx, res in enumerate(results, start=2):
        conf = res.get("CONFIANZA", "BAJA")
        for col_offset, col_name in enumerate(new_cols):
            cell = ws.cell(row=row_idx,
                           column=start_col + col_offset,
                           value=res.get(col_name, ""))
            if col_name == "CONFIANZA" and conf in fills:
                cell.fill = fills[conf]

    wb.save(output_path)
    print(f"  Guardado: {output_path}")

    total = len(results)
    alta  = sum(1 for r in results if r["CONFIANZA"] == "ALTA")
    media = sum(1 for r in results if r["CONFIANZA"] == "MEDIA")
    baja  = sum(1 for r in results if r["CONFIANZA"] == "BAJA")
    rev   = sum(1 for r in results if r["PAIS"] == "REVISAR")

    print(f"\n  RESUMEN:")
    print(f"    Total registros : {total:,}")
    print(f"    Confianza ALTA  : {alta:,}  ({alta/total*100:.1f}%)")
    print(f"    Confianza MEDIA : {media:,}  ({media/total*100:.1f}%)")
    print(f"    Confianza BAJA  : {baja:,}  ({baja/total*100:.1f}%)")
    print(f"    Requieren revisión (PAIS=REVISAR): {rev:,}  ({rev/total*100:.1f}%)")


# ═══════════════════════════════════════════════════════════════════════
# 9. DEMO / TESTING
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    test_cases = [
        ("Akal-C",                                              "México",   "Campeche"),
        ("Dos Bocas Terminal",                                  "México",   "Tabasco"),
        ("Salina Cruz Refinery",                                "México",   "Oaxaca"),
        ("Cd. del Carmen Offshore",                             "México",   "Campeche"),
        ("COATZACOALCOS, VER.",                                 "México",   "Veracruz"),
        ("Coatzacoalcos - Veracruz",                            "México",   "Veracruz"),
        ("coatzacolacos villahermosa",                          "México",   "Veracruz"),
        ("CARR. MEX-TUXPAN KM. 95, TULA DE ALLENDE, HIDALGO.", "México",   "Hidalgo"),
        ("CARR. MONTERREY-CD VICTORIA 85, MONTERREY, NUEVO LEON.", "México","Nuevo León"),
        ("AGUAS DEL GOLFO DE MEXICO, REGION MARINA NOROESTE",  "México",   "Campeche"),
        ("Chalan Pemex 538",                                    "México",   "Campeche"),
        ("BALBOA, PANAMA",                                      "Panamá",   ""),
        ("BOLIVAR ROADS, GALVESTON, TEXAS, USA",                "EUA",      "Texas"),
        ("Amberes, Belgica",                                    "Bélgica",  "Amberes"),
        ("2G73+GJH MINATITLAN, VERACRUZ.",                      "México",   "Veracruz"),
        ("Cuidad Madero, Tamaulipas",                           "México",   "Tamaulipas"),
        ("Cosoloacaque, Veracruz",                              "México",   "Veracruz"),
        ("AEROPUERTO DE SAN JOSE DEL CABO CONOCIDA",           "México",   "Baja California Sur"),
        ("CDMX AZCAPOTZALCO",                                   "México",   "Ciudad de México"),
        ("BCS",                                                  "México",   "Baja California Sur"),
        ("DE LA TAD SALINA CRUZ, OAX., HACIA LA TAD VILLAHERMOSA, TAB.", "México", "Oaxaca"),
        ("DURANTE SU TRAYECTO DE LA TAR PAJARITOS, VERACRUZ, HACIA LA TAR TOLUCA, EDO DE MEXICO",
         "México", "Veracruz"),
        ("Banda oeste del muelle de la terminal maritima de pemex en el puerto de Veracruz.",
         "México", "Veracruz"),
        ("-",                                                   "REVISAR",  "REVISAR"),
        ("CD",                                                  "REVISAR",  "REVISAR"),
    ]

    print("=" * 90)
    print(f"  {'LOCATION':<50} {'ESPERADO':^20} {'OBTENIDO':^12} {'OK'}")
    print("=" * 90)

    correct = 0
    for loc, exp_pais, exp_estado in test_cases:
        res = classify_location(loc, verbose=False)
        ok = "✓" if res["PAIS"] == exp_pais else "✗"
        if res["PAIS"] == exp_pais:
            correct += 1
        esperado = f"{exp_pais}/{exp_estado or '---'}"
        obtenido = f"{res['PAIS']}/{res['ESTADO'] or '---'}"
        flag = f"[{res['CONFIANZA']}]"
        print(f"  {loc[:49]:<50} {esperado:^20} {obtenido:^28} {flag} {ok}")

    print("=" * 90)
    print(f"  Precision: {correct}/{len(test_cases)} ({correct/len(test_cases)*100:.0f}%)")

    # Procesar Excel
    import os
    input_file  = "/sessions/sweet-practical-goodall/mnt/uploads/Location_Ejemplos.xlsx"
    output_file = "/sessions/sweet-practical-goodall/mnt/Codigos Transporte/Location_Clasificado.xlsx"

    if os.path.exists(input_file):
        print(f"\nProcesando: {input_file}")
        process_excel(input_file, output_file, verbose=True)
